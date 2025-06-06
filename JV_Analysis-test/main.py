__author__ = "Edgar Nandayapa"
__version__ = "v0.0.1 2023"

from glob import glob
import pandas as pd
import seaborn as sns
import operator
import os
import re
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
# from openpyxl.styles import Font
import warnings
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.express as px

warnings.filterwarnings("ignore", message=".*cannot be placed.*", category=UserWarning)

def find_and_list_files(folder_path):
    file_patterns = ["JV_*.txt", "**/*JV_*.csv", "**/*JV_*.txt"]
    file_list = []
    for pattern in file_patterns:
        file_list.extend(glob(os.path.join(folder_path, pattern), recursive=True))
    file_list = list(set(file_list))

    return file_list


def load_files(file_list):
    # Consolidate file patterns for searching
    file_list.sort(key=natural_keys)
    # file_list = sorted(file_list, key=lambda x: int(x.split('\\')[-1].replace('JV_', '').replace('.txt', '')))

    # Initialize empty DataFrames for merged results
    jv_chars_merged = pd.DataFrame()
    curves_merged = pd.DataFrame()

    # Process each file
    for file_path in file_list:
        try:
            # Extract JV Characteristics and JV curve from the file
            jv_chars, jv_curve = process_file(file_path)
            # Merge data into cumulative DataFrames
            jv_chars_merged = pd.concat([jv_chars_merged, jv_chars],
                                        ignore_index=True) if not jv_chars.empty else jv_chars_merged
            curves_merged = pd.concat([curves_merged, jv_curve]) if not jv_curve.empty else curves_merged

        except Exception as e:  # Catch all exceptions to avoid stopping the loop
            print(f"Error processing {file_path}: {e}")
    curves_merged = curves_merged.reset_index()
    # Check if data was successfully loaded
    if jv_chars_merged.empty and curves_merged.empty:
        print("One of the files has an issue.")

    return jv_chars_merged, curves_merged


def load_data_from_paths(file_paths):
    """
    Simplified data loading function for Jupyter use
    
    Parameters:
    -----------
    file_paths : list
        List of file paths to process
        
    Returns:
    --------
    tuple : (jv_characteristics_df, jv_curves_df)
    """
    if isinstance(file_paths, str):
        file_paths = [file_paths]
    
    return load_files(file_paths)


def load_data_from_folder(folder_path):
    """
    Load all JV files from a folder
    
    Parameters:
    -----------
    folder_path : str
        Path to folder containing JV files
        
    Returns:
    --------
    tuple : (jv_characteristics_df, jv_curves_df)
    """
    file_list = find_and_list_files(folder_path)
    return load_files(file_list)


def replace_current_density_unit(idx):
    # This regular expression matches (mA/cm²) or (mA/cm^2) and captures the "mA/cm" part before the ² or ^2
    pattern = r'\(mA/cm(?:²|\^2)\)'
    replacement = '(mA/cm2)'
    return re.sub(pattern, replacement, idx)


def process_file(file_path):
    # Determines delimiter based on file extension
    linepos = find_separators_in_file(file_path)
    delimiter = '\t' if file_path.endswith('.txt') else ','

    try:
        # Initial attempt to read JV Characteristics
        jv_chars = pd.read_csv(file_path, skiprows=linepos[0], header=0, index_col=0, nrows=9,
                               delimiter=delimiter).transpose()
        # Attempt to read JV Curve - adjust parameters as per your file structure
        jv_curve = pd.read_csv(file_path, skiprows=linepos[1], header=0, index_col=None,
                               delimiter=delimiter).transpose()

        # Replace problematic character
        jv_chars.columns = [col.replace('²', '2') for col in jv_chars.columns]
        jv_curve.index = [replace_current_density_unit(idx) for idx in jv_curve.index]

        if not jv_chars.empty:
            jv_chars = add_extra_info(jv_chars, file_path, data_type='chars')

        if not jv_curve.empty:
            jv_curve = add_extra_info(jv_curve, file_path, data_type='curve')

    except pd.errors.EmptyDataError:
        jv_chars = pd.DataFrame()
        jv_curve = pd.DataFrame()

    return jv_chars, jv_curve


def add_extra_info(df, file_path, data_type):
    """
    Adds extra information to the DataFrame based on file path and data type.
    """
    import re
    
    # ADD THESE DEBUG LINES TEMPORARILY
    print(f"DEBUG - file_path: {file_path}")
    print(f"DEBUG - basename: {os.path.basename(file_path)}")
    filename = os.path.basename(file_path)
    status_match = re.search(r'[LD]\d+', filename)
    print(f"DEBUG - status_match: {status_match.group(0) if status_match else 'No match'}")
    
    if not status_match:
        # Try alternative pattern
        status_match = re.search(r'([LD]\d+)', filename)
    
    status = status_match.group(1) if status_match else 'N/A'
    
    # Original sample extraction
    sample_name = file_path.split("JV_")[-1].rsplit(".", 1)[0]
    
    # Add status info to sample name for clarity
    if status != 'N/A':
        sample_name_with_status = f"{sample_name}_{status}"
    else:
        sample_name_with_status = sample_name
    
    df['sample'] = sample_name_with_status  # Include status in sample name
    df['original_sample'] = sample_name      # Keep original sample name
    df['batch'] = norm_path.split(os.sep)[-2]
    df['condition'] = pd.NA
    df['status'] = status
    
    # Debug print
    print(f"File: {filename} -> Sample: {sample_name_with_status}, Status: {status}")

    split_index = df.index.to_series().str.split('_', expand=True)
    if data_type == "chars":
        df[['cell', 'direction', 'ilum']] = split_index

    if data_type == 'curve':
        df[['variable', 'cell', 'direction', 'ilum']] = split_index

    return df


def find_separators_in_file(file_path):
    with open(file_path, "r") as file:
        lines = file.readlines()

    positions = []
    for index, line in enumerate(lines):
        if line.strip() == "--":
            positions.append(index + 1)
            # print(f"'--' found at line {index + 1}")

    return positions


def atoi(text):
    return int(text) if text.isdigit() else text


def natural_keys(text):
    return [atoi(c) for c in re.split(r'(\d+)', text)]


def name_by_condition(data, key_list, value_list):
    condition_dict = dict(zip(key_list, value_list))

    data["condition"] = data["sample"].map(condition_dict)

    return data


def data_filter_setup(df, filter_list):
    # Filter conditions
    # par = ["PCE(%)", "FF(%)", "FF(%)", "Voc(V)", "Jsc(mA/cm2)", "ilum"]
    # ope = ["<", "<", ">", "<", ">", "=="]
    # val = [40, 89, 24, 2, -30, "Light"]
    if not filter_list:
        filter_list = [("PCE(%)", "<", "40"), ("FF(%)", "<", "89"), ("FF(%)", ">", "24"), ("Voc(V)", "<", "2"),
                       ("Jsc(mA/cm2)", ">", "-30")]

    # List of operators
    operat = {"<": operator.lt, ">": operator.gt, "==": operator.eq,
              "<=": operator.le, ">=": operator.ge, "!=": operator.ne}

    data = df.copy()

    # Initialize the filter_reason column with empty strings
    data['filter_reason'] = ''
    filtering_options = []
    # for col, op, va in zip(par, ope, val):
    for col, op, va in filter_list:
        # Update the filter_reason for rows that do not meet the condition
        #mask = operat[op](data[col], float(va))
        mask = ~operat[op](data[col], float(va))
        #data.loc[~mask, 'filter_reason'] += f'{col} {op} {va}, '
        data.loc[mask, 'filter_reason'] += f'{col} {op} {va}, '
        filtering_options.append(f'{col} {op} {va}')

    # Filter out rows that have any filter_reason
    trash = data[data['filter_reason'] != ''].copy()
    # Remove rows from data that were moved to trash
    data = data[data['filter_reason'] == '']
    # Clean up the filter_reason string by removing the trailing comma and space
    trash['filter_reason'] = trash['filter_reason'].str.rstrip(', ')

    print(f"\n {trash.shape[0]} of {df.shape[0]} samples were removed based on the specified filters: "
          f"{',  '.join(filtering_options)}.\n")
    print(trash[['sample', 'cell', 'filter_reason']].to_string(index=False))

    return data, trash, filtering_options


def jv_plot_curve_best(jvc, cur, show_grid=True, show_annotations=True, show_device_label=True):
    """Plot the JV curve of the best device using Plotly with toggleable elements
    
    Parameters:
    -----------
    jvc : DataFrame
        JV characteristics data
    cur : DataFrame
        JV curve data
    show_grid : bool, default=True
        Whether to show grid lines
    show_annotations : bool, default=True
        Whether to show performance parameter annotations
    show_device_label : bool, default=True
        Whether to show the device identification label
    """
    import plotly.graph_objects as go
    
    # Find best device
    best_idx = jvc["PCE(%)"].idxmax()
    best_device = jvc.loc[best_idx]
    sample, cell = best_device["sample"], best_device["cell"]
    
    # Get data for best device
    device_data = cur.loc[(cur["sample"] == sample) & (cur["cell"] == cell)]
    
    # Handle alternative identifier if needed
    if len(device_data) == 0:
        sample = best_device["identifier"]
        device_data = cur.loc[(cur["sample"] == sample) & (cur["cell"] == cell)]
    
    # Prepare data for plotting
    plot_data = (device_data.drop(columns=["index", "sample", "cell", "direction", 
                                          "ilum", "batch", "condition"], errors='ignore')
                           .set_index("variable").T)
    
    directions = device_data.loc[device_data["variable"] == "Voltage (V)", "direction"].values
    illuminations = device_data.loc[device_data["variable"] == "Voltage (V)", "ilum"].values
    
    # Create figure
    fig = go.Figure()
    
    # Add axis lines at origin
    fig.add_hline(y=0, line=dict(color="gray", width=2))
    fig.add_vline(x=0, line=dict(color="gray", width=2))
    
    # Plot JV curves for each direction
    for i, direction in enumerate(directions):
        x_data = plot_data["Voltage (V)"].iloc[:, i]
        y_data = plot_data["Current Density(mA/cm2)"].iloc[:, i]
        
        fig.add_trace(go.Scatter(
            x=x_data, y=y_data,
            mode='lines+markers',
            marker_symbol='x' if direction == "Reverse" else 'circle',
            name=f"{direction} ({illuminations[i]})",
            hovertemplate='V: %{x:.3f} V<br>J: %{y:.3f} mA/cm²<extra></extra>'
        ))
    
    # Add MPP points
    for direction in ["Forward", "Reverse"]:
        device_params = jvc.loc[(jvc["sample"] == sample) & 
                               (jvc["cell"] == cell) & 
                               (jvc["direction"] == direction)]
        if not device_params.empty:
            v_mpp = device_params[f'V_mpp(V)'].iloc[0]
            j_mpp = device_params[f'J_mpp(mA/cm2)'].iloc[0]
            
            fig.add_trace(go.Scatter(
                x=[v_mpp], y=[j_mpp],
                mode='markers',
                marker=dict(color='red', size=10, 
                           symbol='x' if direction == "Reverse" else 'circle'),
                name=f'{direction} MPP',
                hovertemplate=f'MPP {direction}<br>V: {v_mpp:.3f} V<br>J: {j_mpp:.3f} mA/cm²<extra></extra>'
            ))
    
    # Add performance parameters as annotations (conditionally)
    if show_annotations:
        # Reverse direction parameters
        params_rev = jvc.loc[(jvc["sample"] == sample) & 
                            (jvc["cell"] == cell) & 
                            (jvc["direction"] == "Reverse")]
        if not params_rev.empty:
            p_rev = params_rev.iloc[0]
            text_rev = f"""        Rev:
<br>Voc: {p_rev['Voc(V)']:>5.2f}
<br>Jsc:  {p_rev['Jsc(mA/cm2)']:>5.1f}
<br>FF:   {p_rev['FF(%)']:>5.1f}
<br>PCE: {p_rev['PCE(%)']:>5.1f}"""
            
            fig.add_annotation(x=0.24, y=-5, text=text_rev, showarrow=False,
                              font=dict(size=12), align="left")
        
        # Forward direction parameters
        params_for = jvc.loc[(jvc["sample"] == sample) & 
                            (jvc["cell"] == cell) & 
                            (jvc["direction"] == "Forward")]
        if not params_for.empty:
            p_for = params_for.iloc[0]
            text_for = f"For:<br>{p_for['Voc(V)']:.2f} V<br>{p_for['Jsc(mA/cm2)']:.1f} mA/cm²<br>{p_for['FF(%)']:.1f}%<br>{p_for['PCE(%)']:.1f}%"
            
            fig.add_annotation(x=0.55, y=-5, text=text_for, showarrow=False,
                              font=dict(size=12), align="left")
    
    # Add device label (conditionally)
    if show_device_label:
        fig.add_annotation(x=0.20, y=1.5, 
                          text=f"Sample: {sample} ({cell})",
                          showarrow=False, font=dict(size=13), align="left")
    
    # Configure layout with conditional grid
    grid_config = dict(showgrid=show_grid, gridcolor='lightgray') if show_grid else dict(showgrid=False)
    
    fig.update_layout(
        title="JV Curve - Best Device",
        xaxis_title='Voltage [V]',
        yaxis_title='Current Density [mA/cm²]',
        xaxis=dict(range=[-0.2, 1.35], **grid_config),
        yaxis=dict(range=[-25, 3], **grid_config),
        template="plotly_white",
        legend=dict(x=1.02, y=1, xanchor="left", yanchor="top"),
        autosize=True,
        margin=dict(autoexpand=True)
    )
    
    # Save if not in Jupyter
    filename = "JV_best_device.html"

    fig.update_layout(
    autosize=True,
    margin=dict(autoexpand=True)
    )
    
    return fig, filename

def jv_plot_together(df1, df2, namestring):
    """Plot JV curves together using Plotly with full interactivity"""
    
    # Prepare the data frame
    if namestring == "All":
        df2_copy = df2.copy()
    else:
        # Filter curves to match JV data combinations
        matching_cols = ['sample', 'batch', 'cell', 'direction', 'ilum']
        jv_combinations = df1[matching_cols].drop_duplicates()
        df2_copy = pd.merge(jv_combinations, df2, on=matching_cols, how='inner')
        
        if df2_copy.empty:
            fig = go.Figure()
            fig.update_layout(title=f"No curves found for {namestring}")
            return fig, f"empty_{namestring}.html"

    # Prepare plotting data
    df2_plot = drop_extra_cols_and_ready_to_plot(df2_copy)

    # Rename columns for plotting
    cols = []
    counters = {'Voltage (V)': 0, 'Current Density(mA/cm2)': 0}
    for col in df2_plot.columns:
        counters[col] += 1
        cols.append(f"{col} {counters[col]}")
    df2_plot.columns = cols

    # Create Plotly figure
    fig = go.Figure()

    # Add axis lines at 0
    fig.add_shape(type="line", x0=-0.2, y0=0, x1=1.35, y1=0, line=dict(color="gray", width=2))
    fig.add_shape(type="line", x0=0, y0=-25, x1=0, y1=3, line=dict(color="gray", width=2))

    # Extract source data for plotting
    required_cols = ['sample', 'batch', 'cell', 'direction', 'ilum', 'variable']
    source_data = df2_copy[required_cols].drop_duplicates()
    source_data = source_data[source_data['variable'] == 'Voltage (V)']
    
    # Get PCE values for lookup
    pce_values = {}
    if 'PCE(%)' in df1.columns:
        for _, row in df1.iterrows():
            key = (row['sample'], row['batch'], row['cell'], row['direction'], row['ilum'])
            pce_values[key] = row['PCE(%)']
    
    # Generate distinct colors
    import random, colorsys
    
    def generate_distinct_colors(n):
        colors = []
        for i in range(n):
            hue = (i * 0.618033988749895) % 1
            saturation = 0.7 + random.random() * 0.3
            value = 0.7 + random.random() * 0.3
            r, g, b = colorsys.hsv_to_rgb(hue, saturation, value)
            colors.append(f"rgb({int(r*255)}, {int(g*255)}, {int(b*255)})")
        return colors
    
    unique_samples = source_data[['sample', 'batch', 'cell']].drop_duplicates()
    all_colors = generate_distinct_colors(len(unique_samples) * 2)
    random.shuffle(all_colors)
    
    added_traces = {}
    num_pairs = len(df2_plot.columns) // 2
    
    # Add traces for each curve
    for i in range(1, num_pairs + 1):
        voltage_col = f'Voltage (V) {i}'
        current_col = f'Current Density(mA/cm2) {i}'
        
        if voltage_col not in df2_plot.columns or current_col not in df2_plot.columns:
            continue
            
        if i <= len(source_data):
            sample_info = source_data.iloc[i-1]
            sample_name = sample_info['sample']
            batch_name = sample_info['batch']
            cell_name = sample_info['cell']
            direction = sample_info['direction']
            ilum = sample_info['ilum']
        else:
            continue
        
        # Get PCE value
        pce_key = (sample_name, batch_name, cell_name, direction, ilum)
        pce_val = pce_values.get(pce_key, None)
        pce_str = f", PCE: {pce_val:.2f}%" if pce_val is not None else ", PCE: N/A"
        
        # Create unique identifier and color assignment
        sample_cell_key = (sample_name, batch_name, cell_name)
        
        if sample_cell_key not in added_traces:
            color_idx = len(added_traces) % len(all_colors)
            added_traces[sample_cell_key] = {
                'color': all_colors[color_idx], 
                'forward_added': False, 'reverse_added': False
            }
        
        is_forward = direction == 'Forward'
        line_dash = 'solid' if is_forward else 'dash'
        
        # Check if this direction was already added
        if (is_forward and added_traces[sample_cell_key]['forward_added']) or \
           (not is_forward and added_traces[sample_cell_key]['reverse_added']):
            continue
        
        # Mark this direction as added
        added_traces[sample_cell_key][f"{'forward' if is_forward else 'reverse'}_added"] = True
        
        # Create trace name
        curve_name = f"Sample {sample_name}, Cell {cell_name} ({direction}, {ilum}{pce_str})"
        
        # Add trace to figure
        fig.add_trace(go.Scatter(
            x=df2_plot[voltage_col], 
            y=df2_plot[current_col],
            mode='lines', 
            line=dict(color=added_traces[sample_cell_key]['color'], width=2, dash=line_dash),
            name=curve_name, 
            showlegend=True,
            hovertemplate=(
                f"Voltage: %{{x:.3f}} V<br>Current Density: %{{y:.3f}} mA/cm²<br>{pce_str}<br>" +
                f"Sample: {sample_name}<br>Cell: {cell_name}<br>Direction: {direction}<br>Illumination: {ilum}<extra></extra>"
            )
        ))
    
    # Update layout
    fig.update_layout(
        title=f'{namestring} J-V Curves ({len(added_traces)} combinations)',
        xaxis_title='Voltage [V]', 
        yaxis_title='Current Density [mA/cm²]',
        xaxis=dict(range=[-0.2, 1.35]), 
        yaxis=dict(range=[-25, 3]),
        template="plotly_white", 
        height=700,
        legend=dict(x=1.02, y=1, xanchor='left', yanchor='top', font={'size': 10}),
        margin=dict(l=80, r=200, t=100, b=80)
    )
    
    # Add grid lines
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')
    
    # Save figure
    image_name = f"JV_together_{namestring}.html"
        
    fig.update_layout(
    autosize=True,
    margin=dict(autoexpand=True)
    )
    
    return fig, image_name

def jv_plot_by_cell_3x2(df, sample):
    """Plot JV curves for each cell in a 2x3 grid using Plotly"""
    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    
    # Filter the DataFrame for the specified sample
    focus = df[df["sample"] == sample]

    # Group the DataFrame by 'cell'
    grouped = focus.groupby('cell')
    
    # Create 2x3 subplot grid
    fig = make_subplots(rows=2, cols=3, 
                        subplot_titles=[f"Cell {cell}" for cell in grouped.groups.keys()],
                        shared_xaxes=True, shared_yaxes=True)

    # Dynamically identify measurement columns
    # These are numeric columns that aren't metadata
    metadata_cols = ["index", "sample", "cell", "direction", "ilum", "batch", "condition", "variable"]
    all_cols = df.columns.tolist()
    measurement_cols = [col for col in all_cols if col not in metadata_cols and col != '']
    
    # Iterate through each cell
    for i, (cell, group) in enumerate(grouped):
        row = i // 3 + 1  # Calculate row (1-indexed)
        col = i % 3 + 1   # Calculate column (1-indexed)
        
        if i >= 6:  # Check to prevent index error if there are more than 6 cells
            break

        # Add x and y axis lines at 0
        fig.add_shape(type="line", x0=-0.2, y0=0, x1=1.35, y1=0, 
                    line=dict(color="gray", width=2), row=row, col=col)
        fig.add_shape(type="line", x0=0, y0=-25, x1=0, y1=3, 
                    line=dict(color="gray", width=2), row=row, col=col)

        # Process each direction and illumination combination
        for direction in group['direction'].unique():
            for ilum in group['ilum'].unique():
                # Get voltage data
                voltage_row = group[
                    (group["direction"] == direction) & 
                    (group["ilum"] == ilum) & 
                    (group["variable"] == "Voltage (V)")
                ]
                
                # Get current density data
                current_row = group[
                    (group["direction"] == direction) & 
                    (group["ilum"] == ilum) & 
                    (group["variable"] == "Current Density(mA/cm2)")
                ]
                
                if voltage_row.empty or current_row.empty:
                    print(f"Debug: No data for Cell {cell}, {direction}, {ilum}")
                    continue
                
                # Extract the measurement values using available columns
                try:
                    voltage_values = voltage_row[measurement_cols].iloc[0].values
                    current_values = current_row[measurement_cols].iloc[0].values

                    
                    # Convert to mA/cm² if values are very small (likely in A/cm²)
                    if abs(current_values.max()) < 0.1:  # If max current is less than 0.1, likely in A
                        current_values_ma = current_values * 1000  # Convert A to mA
                    else:
                        current_values_ma = current_values
                    
                    # Plot the curve
                    label = f"{direction} ({ilum})"
                    
                    fig.add_trace(
                        go.Scatter(
                            x=voltage_values, 
                            y=current_values_ma,
                            mode="lines+markers",
                            name=label,
                            legendgroup=label,
                            showlegend=(i == 0),  # Only show legend for the first cell
                            hovertemplate='Voltage: %{x:.3f} V<br>Current Density: %{y:.3f} mA/cm²<br>Cell: ' + str(cell) + '<br>' + label
                        ),
                        row=row, col=col
                    )                    
                except Exception as e:
                    print(f"Debug: Error processing Cell {cell}, {direction}, {ilum}: {e}")
                    print(f"Debug: Voltage row columns: {voltage_row.columns.tolist()}")
                    continue

    # Update layout
    fig.update_layout(
        title=f"Sample {sample}",
        height=700,
        width=1000,
        template="plotly_white",
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="center",
            x=0.5
        )
    )

    # Update all x and y axes
    for i in range(1, 7):
        if i <= len(grouped):
            row = (i-1) // 3 + 1
            col = (i-1) % 3 + 1
            
            fig.update_xaxes(
                title_text="Voltage [V]" if row == 2 else None,
                range=[-0.2, 1.35],
                showgrid=True, 
                gridwidth=1, 
                gridcolor='lightgray',
                row=row, col=col
            )
            
            fig.update_yaxes(
                title_text="Current Density [mA/cm²]" if col == 1 else None,
                range=[-25, 3],
                showgrid=True, 
                gridwidth=1, 
                gridcolor='lightgray',
                row=row, col=col
            )

    # Save figure if not in Jupyter
    image_name = f"JV_cells_by_sample_{sample}.html"

    fig.update_layout(
    autosize=True,
    margin=dict(autoexpand=True)
    )

    return fig, image_name


def jv_plot_by_substrate(df, sample):
    """Plot JV curves for all cells on a substrate using Plotly"""
    import plotly.graph_objects as go
    
    focus = df[df["sample"] == sample]
    
    # Create Plotly figure
    fig = go.Figure()

    # Add x and y axis lines at 0
    fig.add_shape(type="line", x0=-0.2, y0=0, x1=1.35, y1=0, line=dict(color="gray", width=2))
    fig.add_shape(type="line", x0=0, y0=-25, x1=0, y1=3, line=dict(color="gray", width=2))

    grouped = focus.groupby('cell')
    
    # Create a color scale for different cells
    colors = [f"hsl({i*360/grouped.ngroups},100%,50%)" for i in range(grouped.ngroups)]
    
    # Dynamically identify measurement columns
    metadata_cols = ["index", "sample", "cell", "direction", "ilum", "batch", "condition", "variable"]
    all_cols = df.columns.tolist()
    measurement_cols = [col for col in all_cols if col not in metadata_cols and col != '']

    for i, ((cell, group), color) in enumerate(zip(grouped, colors)):
        # Iterate through directions and illumination conditions
        for direction in group['direction'].unique():
            for ilum in group['ilum'].unique():
                # Get voltage data
                voltage_row = group[
                    (group["direction"] == direction) & 
                    (group["ilum"] == ilum) & 
                    (group["variable"] == "Voltage (V)")
                ]
                
                # Get current density data
                current_row = group[
                    (group["direction"] == direction) & 
                    (group["ilum"] == ilum) & 
                    (group["variable"] == "Current Density(mA/cm2)")
                ]
                
                if voltage_row.empty or current_row.empty:
                    print(f"Debug: No data for Cell {cell}, {direction}, {ilum}")
                    continue
                
                # Extract the measurement values using available columns
                try:
                    voltage_values = voltage_row[measurement_cols].iloc[0].values
                    current_values = current_row[measurement_cols].iloc[0].values
                    
                    # Convert to mA/cm² if values are very small (likely in A/cm²)
                    if abs(current_values.max()) < 0.1:  # If max current is less than 0.1, likely in A
                        current_values_ma = current_values * 1000  # Convert A to mA
                    else:
                        current_values_ma = current_values

                    # Determine line style based on direction
                    line_dash = 'solid' if direction == 'Forward' else 'dash'
                    
                    # Plot the curve
                    label = f"Cell {cell}, {direction} ({ilum})"
                    
                    fig.add_trace(go.Scatter(
                        x=voltage_values, 
                        y=current_values_ma,
                        mode="lines",
                        name=label,
                        line=dict(color=color, dash=line_dash),
                        hovertemplate='Voltage: %{x:.3f} V<br>Current Density: %{y:.3f} mA/cm²<br>' + label
                    ))
                    
                except Exception as e:
                    print(f"Debug: Error processing Cell {cell}, {direction}, {ilum}: {e}")
                    print(f"Debug: Available columns in voltage_row: {voltage_row.columns.tolist()}")
                    continue

    # Update layout
    fig.update_layout(
        title=f"Sample {sample}",
        xaxis_title='Voltage [V]',
        yaxis_title='Current Density [mA/cm²]',
        xaxis=dict(range=[-0.2, 1.35]),
        yaxis=dict(range=[-25, 3]),
        template="plotly_white",
        legend=dict(
            x=1.02,
            y=1,
            xanchor="left",
            yanchor="top"
        )
    )

    # Add grid lines
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')

    # Save figure if not in Jupyter
    image_name = f"JV_combined_sample_{sample}.html"

    fig.update_layout(
    autosize=True,
    margin=dict(autoexpand=True)
    )

    return fig, image_name
    

def drop_extra_cols_and_ready_to_plot(df):
    cols_to_remove = ["index", "sample", "cell", "direction", "ilum", "batch", "condition"]

    common_cols_to_remove = df.columns.intersection(cols_to_remove)
    df_clean = df.copy().drop(columns=common_cols_to_remove)

    df_clean = df_clean.set_index(["variable"]).T
    return df_clean


def boxplot_all_cells(data, var_x, var_y, filtered_info, datatype):
    """Create a boxplot with all cells using Plotly"""
    names_dict = {
        "voc": 'Voc(V)', "jsc": 'Jsc(mA/cm2)', "ff": 'FF(%)', "pce": 'PCE(%)',
        "vmpp": 'V_mpp(V)', "jmpp": 'J_mpp(mA/cm2)', "pmpp": 'P_mpp(mW/cm2)',
        "rser": 'R_series(Ohmcm2)', "rshu": 'R_shunt(Ohmcm2)'
    }
    var_name_y = names_dict[var_y]
    trash, filters = filtered_info

    # Basic data preparation
    try:
        data["sample"] = data["sample"].astype(int)
    except ValueError:
        pass
    data['Jsc(mA/cm2)'] = data['Jsc(mA/cm2)'].abs()

    # Get categories
    orderc = data.groupby(var_x)[var_name_y].describe().sort_index()["count"].index
    data_counts = data.groupby(var_x)[var_name_y].count().to_dict()
    trash_counts = trash.groupby(var_x)[var_name_y].count().to_dict() if not trash.empty else {}

    # Create figure
    fig = go.Figure()
    
    # Color palette
    colors = [
        'rgba(93, 164, 214, 0.7)', 'rgba(255, 144, 14, 0.7)', 
        'rgba(44, 160, 101, 0.7)', 'rgba(255, 65, 54, 0.7)', 
        'rgba(207, 114, 255, 0.7)', 'rgba(127, 96, 0, 0.7)',
        'rgba(255, 140, 184, 0.7)', 'rgba(79, 90, 117, 0.7)',
        'rgba(222, 158, 54, 0.7)', 'rgba(82, 182, 133, 0.7)',
        'rgba(148, 103, 189, 0.7)', 'rgba(23, 190, 207, 0.7)'
    ]
    
    # Create boxplots
    for i, category in enumerate(orderc):
        category_data = data[data[var_x] == category][var_name_y].dropna()
        if not category_data.empty:
            data_count = data_counts.get(category, 0)
            trash_count = trash_counts.get(category, 0)
            
            category_name = (f"{category} (n={data_count})" if trash_count == 0 
                           else f"{category} ({data_count}/{data_count + trash_count})")
            
            fig.add_trace(go.Box(
                y=category_data,
                name=category_name,
                boxpoints='all',
                pointpos=0,
                jitter=0.5,
                marker=dict(size=5, opacity=0.7, color='rgba(0,0,0,0.7)'),
                line=dict(width=1.5),
                fillcolor=colors[i % len(colors)],
                boxmean=True,
                hovertemplate=(f"<b>{category}</b><br>Value: %{{y:.3f}}<br>Count: {data_count}")
            ))
    
    # Layout
    title_text = f"Boxplot of {var_y} by {var_x}"
    if datatype == "junk":
        title_text += " (filtered out)"
    
    fig.update_layout(
        title={'text': title_text, 'x': 0.5, 'font': dict(size=18)},
        xaxis_title=var_x,
        yaxis_title=var_name_y,
        boxmode='group',
        boxgap=0.05,
        template="plotly_white",
        margin=dict(l=40, r=40, t=80, b=60),
        showlegend=False
    )
    
    # Make boxes wide
    fig.update_traces(width=0.8, quartilemethod="linear")
    
    # Rotate labels if many categories
    if len(orderc) > 4:
        fig.update_layout(xaxis=dict(tickangle=-10, tickfont=dict(size=10)))

    sample_name = f"boxplot{'j' if datatype == 'junk' else ''}_{var_y}_by_{var_x}.html"

    fig.update_layout(
    autosize=True,
    margin=dict(autoexpand=True)
    )

    return fig, sample_name


def boxplot_paired_by_direction(data, var_x, var_y, filtered_info, datatype):
    """Create a paired boxplot with forward/backward direction using different colors"""
    names_dict = {
        "voc": 'Voc(V)', "jsc": 'Jsc(mA/cm2)', "ff": 'FF(%)', "pce": 'PCE(%)',
        "vmpp": 'V_mpp(V)', "jmpp": 'J_mpp(mA/cm2)', "pmpp": 'P_mpp(mW/cm2)',
        "rser": 'R_series(Ohmcm2)', "rshu": 'R_shunt(Ohmcm2)'
    }
    var_name_y = names_dict[var_y]
    trash, filters = filtered_info

    # Basic data preparation
    try:
        data["sample"] = data["sample"].astype(int)
    except ValueError:
        pass
    data['Jsc(mA/cm2)'] = data['Jsc(mA/cm2)'].abs()

    # Check direction column
    if 'direction' not in data.columns:
        print("Warning: 'direction' column not found in data. Creating dummy direction column.")
        data['direction'] = 'forward'

    # Get categories and directions
    orderc = data.groupby(var_x)[var_name_y].describe().sort_index()["count"].index
    directions = sorted(data['direction'].unique())
    
    # Setup colors and legend tracking
    direction_colors = {}
    legend_added = {}
    for direction in directions:
        direction_lower = direction.lower()
        if 'forward' in direction_lower or 'fwd' in direction_lower:
            direction_colors[direction] = 'rgba(93, 164, 214, 0.8)'   # Blue
        elif 'backward' in direction_lower or 'reverse' in direction_lower or 'rev' in direction_lower:
            direction_colors[direction] = 'rgba(255, 100, 100, 0.8)'   # Red
        else:
            direction_colors[direction] = 'rgba(128, 128, 128, 0.8)'   # Gray
        legend_added[direction] = False

    # Create figure
    fig = go.Figure()
    data_counts = data.groupby([var_x, 'direction'])[var_name_y].count().to_dict()
    
    # Box width setting
    box_width = 0.25  # Adjust this value to make boxes thinner/wider
    
    # Create paired boxplots
    for i, category in enumerate(orderc):
        for j, direction in enumerate(directions):
            category_data = data[(data[var_x] == category) & (data['direction'] == direction)][var_name_y].dropna()
            
            if not category_data.empty:
                data_count = data_counts.get((category, direction), 0)
                show_legend = not legend_added[direction]
                
                if show_legend:
                    legend_added[direction] = True
                    legend_name = direction
                else:
                    legend_name = None
                
                # Calculate position
                offset = (j - (len(directions) - 1) / 2) * box_width
                x_position = i + offset
                
                fig.add_trace(go.Box(
                    y=category_data,
                    name=legend_name,
                    legendgroup=direction,
                    showlegend=show_legend,
                    boxpoints='all',
                    pointpos=0,
                    jitter=0.3,
                    marker=dict(size=4, opacity=0.6, color='rgba(0,0,0,1)'),
                    line=dict(width=2, color=direction_colors[direction]),
                    fillcolor=direction_colors[direction],
                    boxmean=True,
                    width=box_width-0.08,
                    x=[x_position] * len(category_data),
                    hovertemplate=(f"<b>{category} ({direction})</b><br>"
                                 f"Value: %{{y:.3f}}<br>Count: {data_count}")
                ))

    # Layout
    title_text = f"Paired Boxplot of {var_y} by {var_x} (Forward vs Backward)"
    if datatype == "junk":
        title_text += " (filtered out)"
    
    fig.update_layout(
        title={'text': title_text, 'x': 0.5, 'font': dict(size=18)},
        xaxis=dict(tickmode='array', tickvals=list(range(len(orderc))), 
                  ticktext=list(orderc), title=var_x),
        yaxis_title=var_name_y,
        template="plotly_white",
        margin=dict(l=40, r=120, t=80, b=60),
        showlegend=True,
        legend=dict(orientation="v", yanchor="top", y=1, xanchor="left", x=1.02,
                   bgcolor="rgba(255,255,255,0.8)", bordercolor="rgba(0,0,0,0.2)", borderwidth=1)
    )
    
    sample_name = f"boxplot_paired_{'j_' if datatype == 'junk' else ''}{var_y}_by_{var_x}.html"

    fig.update_layout(
    autosize=True,
    margin=dict(autoexpand=True)
    )

    return fig, sample_name


def create_workbook_with_data(data):
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    
    # Write the DataFrame to a sheet named 'All_data'
    ws = wb.create_sheet(title='All_data')
    for r in dataframe_to_rows(data, index=True, header=True):
        ws.append(r)
    
    return wb


def save_combined_excel_data(wb, data, filtered_info, var_x, name_y, var_y, other_df):
    """Save data to Excel workbook - same functionality as the original but returning the workbook"""
    trash, filters = filtered_info

    # Create a new sheet name based on var_x and var_y
    sheet_title = f"{var_y}-by-{var_x}"

    # Check if the sheet already exists
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    # Insert personalized string before the first DataFrame
    ws.append([f"Contents of boxplot for {var_y} by {var_x}"])
    ws.append([])  # Add an empty row for spacing

    # Process and append data and other_df as before
    combined_data = data.copy()
    combined_data['_index'] = combined_data.groupby(var_x).cumcount()
    pivot_table = combined_data.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

    for r in dataframe_to_rows(pivot_table, index=True, header=True):
        ws.append(r)

    # Calculate starting row for the second personalized string
    # It's the current number of rows plus 2 for spacing
    next_row = ws.max_row + 3

    # Insert personalized string before the second DataFrame
    ws.cell(row=next_row, column=1, value="Statistical summary")
    ws.append([])  # Add an empty row for spacing

    for r in dataframe_to_rows(other_df.T, index=True, header=True):
        ws.append(r)

    next_row = ws.max_row + 3

    # Insert personalized string before the second DataFrame
    ws.cell(row=next_row, column=1, value="This is the filtered data")
    ws.append([])  # Add an empty row for spacing

    combined_trash = trash.copy()
    combined_trash['_index'] = combined_trash.groupby(var_x).cumcount()
    pivot_table_trash = combined_trash.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

    # Add rows from the second DataFrame (pivot table)
    for r in dataframe_to_rows(pivot_table_trash, index=True, header=True):
        ws.append(r)

    next_row = ws.max_row + 3

    # Insert personalized string before the second DataFrame
    filter_words = ["Only data within these limits is shown:"] + filters
    for cc, strings in enumerate(filter_words):
        ws.cell(row=next_row + cc, column=1, value=strings)
    ws.append([])  # Add an empty row for spacing

    return wb

def histogram(df, var_y):
    """Create a histogram using Plotly"""
    names_dict = {
        'voc': 'Voc(V)', 'jsc': 'Jsc(mA/cm2)', 'ff': 'FF(%)', 'pce': 'PCE(%)',
        'vmpp': 'V_mpp(V)', 'jmpp': 'J_mpp(mA/cm2)', 'pmpp': 'P_mpp(mW/cm2)',
        'rser': 'R_series(Ohmcm2)', 'rshu': 'R_shunt(Ohmcm2)'
    }

    pl_y = names_dict[var_y]

    # Determine number of bins based on variable
    if var_y == "voc":
        bins = 20
    elif var_y == "jsc":
        bins = 30
    else:
        bins = 40

    # Create the histogram figure
    fig = go.Figure()
    
    # Add the histogram trace
    fig.add_trace(go.Histogram(
        x=df[pl_y],
        nbinsx=bins,
        marker=dict(
            color='rgba(0, 0, 255, 0.6)',
            line=dict(color='rgba(0, 0, 255, 1)', width=1)
        ),
        hovertemplate=f'{pl_y}: %{{x:.3f}}<br>Count: %{{y}}<extra></extra>'
    ))
    
    # Add a line trace for the kernel density estimate if enough data points
    if len(df) > 5:
        from scipy import stats
        kde_x = np.linspace(df[pl_y].min(), df[pl_y].max(), 100)
        kde = stats.gaussian_kde(df[pl_y].dropna())
        kde_y = kde(kde_x) * len(df) * (df[pl_y].max() - df[pl_y].min()) / bins
        
        fig.add_trace(go.Scatter(
            x=kde_x,
            y=kde_y,
            mode='lines',
            line=dict(color='red', width=2),
            name='KDE',
            hoverinfo='skip'
        ))
    
    # Add statistics annotations
    mean_val = df[pl_y].mean()
    median_val = df[pl_y].median()
    std_val = df[pl_y].std()
    
    stats_text = (
        f"Mean: {mean_val:.3f}<br>"
        f"Median: {median_val:.3f}<br>"
        f"Std Dev: {std_val:.3f}<br>"
        f"Count: {len(df)}"
    )
    
    fig.add_annotation(
        x=0.95,
        y=0.95,
        xref="paper",
        yref="paper",
        text=stats_text,
        showarrow=False,
        font=dict(size=12),
        align="right",
        bgcolor="rgba(255, 255, 255, 0.8)",
        bordercolor="black",
        borderwidth=1,
        borderpad=4
    )
    
    # Update layout
    fig.update_layout(
        title=f"Histogram of {pl_y}",
        xaxis_title=pl_y,
        yaxis_title="Frequency",
        template="plotly_white",
        bargap=0.1,
        hovermode='closest'
    )
    
    # Add grid lines
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')

    # Save figure if not in Jupyter
    sample_name = f"histogram_{var_y}.html"

    return fig, sample_name


def plot_list_from_voila(plot_list):
    jvc_dict = {'Voc': 'v', 'Jsc': 'j', 'FF': 'f', 'PCE': 'p', 'R_ser': 'r', 'R_shu': 'h', 'V_mpp': 'u',
                'J_mpp': 'i', 'P_mpp': 'm'}
    box_dict = {'by Batch': 'e', 'by Variable': 'g', 'by Sample': 'a', 'by Cell': 'b', 'by Scan Direction': 'c', 'by Status': 's'}
    cur_dict = {'All cells': 'Cy', 'Only working cells': 'Cz', 'Only not working cells': 'Co', 'Best device only': 'Cw',
                'Separated by cell': 'Cx', 'Separated by substrate': 'Cd'}
    
    # Add option2 variations for Best device only
    best_device_dict = {
        'All annotations': 'Cw',
        'No grid': 'Cw1',
        'No performance': 'Cw2', 
        'No grid and no performance': 'Cw3'
    }
    
    new_list = []
    for plot in plot_list:
        code = ''
        if "Paired Boxplot" in plot[0]:
            code += "P"
            code += (jvc_dict[plot[1]])  # This looks up the variable (Voc, PCE, etc.)
            code += (box_dict[plot[2]])  # This looks up the grouping (by Status, by Sample, etc.)
        elif "omitted" in plot[0]:
            code += "J"
            code += (jvc_dict[plot[1]])  # Variable
            code += (box_dict[plot[2]])  # Grouping
        elif "Boxplot" in plot[0]:
            code += "B"
            code += (jvc_dict[plot[1]])  # Variable  
            code += (box_dict[plot[2]])  # Grouping
        elif "Histogram" in plot[0]:
            code += "H"
            code += (jvc_dict[plot[1]])  # Variable only
        else:
            # Handle JV Curve plots
            if plot[1] == 'Best device only' and len(plot) > 2:
                code += best_device_dict.get(plot[2], 'Cw')
            else:
                code += (cur_dict[plot[1]])
        new_list.append(code)
    return new_list
    

def create_plots(jv_chars, jv_curves, plot_types=None, filter_params=None, conditions=None):
    """
    Master function to create plots from JV data
    
    Parameters:
    -----------
    jv_chars : DataFrame
        JV characteristics data
    jv_curves : DataFrame
        JV curves data
    plot_types : list, optional
        List of plot codes. If None, creates default plots.
    filter_params : list, optional
        List of filter tuples [(parameter, operator, value), ...]
    conditions : dict, optional
        Dictionary mapping samples to conditions
        
    Returns:
    --------
    tuple : (figure_list, figure_names)
    """
    # Set up data
    if conditions:
        samples = list(conditions.keys())
        condition_values = list(conditions.values())
        jv_chars = name_by_condition(jv_chars, samples, condition_values)
        jv_curves = name_by_condition(jv_curves, samples, condition_values)
        has_conditions = True
    else:
        has_conditions = False
    
    # Apply filters
    if filter_params is None:
        filter_params = [("PCE(%)", "<", "40"), ("FF(%)", "<", "89"), ("FF(%)", ">", "24"), 
                        ("Voc(V)", "<", "2"), ("Jsc(mA/cm2)", ">", "-30")]
    
    filtered_jv, omitted_jv, filter_options = data_filter_setup(jv_chars, filter_params)

    print("Available columns:", filtered_jv.columns.tolist())
    print("Sample of data:")
    print(filtered_jv.head())
    
    # Set default plots if none specified
    if plot_types is None:
        plot_types = ["Bpa", "Bpj", "Bpf", "Bpp", "Hp", "Cy"]
    
    # Get unique samples
    samples = find_unique_values(jv_chars)
    
    # Prepare data structures
    jv_data = (filtered_jv, jv_chars, jv_curves)
    support_data = (omitted_jv, filter_options, has_conditions, samples)
    
    # Generate plots
    return plotting_string_action(plot_types, jv_data, support_data)


def plotting_string_action(plot_codes, jv_data, support_data, is_voila=False):
    """
    Generate plots based on plot codes
    
    Parameters:
    -----------
    plot_codes : list
        List of plot code strings (e.g., ['Bpa', 'Hp', 'Cy'])
    jv_data : tuple
        (filtered_jv, complete_jv, complete_curves)
    support_data : tuple
        (omitted_jv, filter_params, has_conditions, samples)
    is_voila : bool, default=False
        Whether called from Voila interface
        
    Returns:
    --------
    tuple : (figure_list, figure_names)
    """
    
    filtered_jv, complete_jv, complete_cur = jv_data
    omitted_jv, filter_pars, is_conditions, _, samples = support_data
    if is_voila:
        plot_codes = plot_list_from_voila(plot_codes)
    
    varx_dict = {"a": "sample", "b": "cell", "c": "direction", "d": "ilum", "e": "batch", "g": "condition", "s": "status"}
    vary_dict = {"v": "voc", "j": "jsc", "f": "ff", "p": "pce", "u": "vmpp", "i": "jmpp", "m": "pmpp", "r": "rser", "h": "rshu", }
    
    fig_list = []
    fig_names = []
    for pl in plot_codes:
        if "g" in pl and not is_conditions:
            continue
            
        # Find var_x and var_y
        var_x = None
        var_y = None
        for key, value in varx_dict.items():
            if key in pl:
                var_x = value
                break
        for key, value in vary_dict.items():
            if key in pl:
                var_y = value
                break
        
        # Handle different plot types
        if "B" in pl and var_x is not None and var_y is not None:
            fig, fig_name = boxplot_all_cells(filtered_jv, var_x, var_y, [omitted_jv, filter_pars], "data")
        elif "P" in pl and var_x is not None and var_y is not None:
            fig, fig_name = boxplot_paired_by_direction(filtered_jv, var_x, var_y, [omitted_jv, filter_pars], "data")
        elif "J" in pl and var_x is not None and var_y is not None:
            fig, fig_name = boxplot_all_cells(omitted_jv, var_x, var_y, [filtered_jv, filter_pars], "junk")
        elif "H" in pl and var_y is not None:
            fig, fig_name = histogram(complete_jv, var_y)
        elif "Cw" in pl:  # Best device variations
            if pl == "Cw1":
                fig, fig_name = jv_plot_curve_best(complete_jv, complete_cur, show_grid=False)
            elif pl == "Cw2":
                fig, fig_name = jv_plot_curve_best(complete_jv, complete_cur, show_annotations=False)
            elif pl == "Cw3":
                fig, fig_name = jv_plot_curve_best(complete_jv, complete_cur, show_grid=False, show_annotations=False)
            else:
                fig, fig_name = jv_plot_curve_best(complete_jv, complete_cur)
        elif "Cc" in pl:
            fig, fig_name = jv_plot_curve_best(complete_jv, complete_cur)
        elif "Cx" in pl:  # Cells per sample
            for s in samples:
                fig, fig_name = jv_plot_by_cell_3x2(complete_cur, s)
                fig_list.append(fig)
                fig_names.append(fig_name)
            continue
        elif "Cd" in pl:  # Cells per substrate
            for s in samples:
                fig, fig_name = jv_plot_by_substrate(complete_cur, s)
                fig_list.append(fig)
                fig_names.append(fig_name)
            continue
        elif "Cy" in pl:  # All data
            fig, fig_name = jv_plot_together(complete_jv, complete_cur, "All")
        elif "Cz" in pl:  # Only filtered (working cells)
            fig, fig_name = jv_plot_together(filtered_jv, complete_cur, "Filtered")
        elif "Co" in pl:  # Only omitted (non-working cells)
            print(f"Plotting ONLY NON-WORKING cells")
            print(f"omitted_jv shape: {omitted_jv.shape}")
            if not omitted_jv.empty:
                print(f"omitted_jv PCE range: {omitted_jv['PCE(%)'].min():.2f} to {omitted_jv['PCE(%)'].max():.2f}")
            if not filtered_jv.empty:
                print(f"filtered_jv PCE range: {filtered_jv['PCE(%)'].min():.2f} to {filtered_jv['PCE(%)'].max():.2f}")
            print(f"sample/cell combinations: {omitted_jv[['sample', 'cell']].drop_duplicates().shape[0]}")
            fig, fig_name = jv_plot_together(omitted_jv, complete_cur, "Omitted")
        else:
            print(f"Command {pl} not recognized")
            continue
            
        fig_list.append(fig)
        fig_names.append(fig_name)
    return fig_list, fig_names


def save_full_data_frame(data):
    """Create a workbook with JV data for Excel export"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    
    # Write the DataFrame to a sheet named 'All_data'
    ws = wb.create_sheet(title='All_data')
    for r in dataframe_to_rows(data, index=True, header=True):
        ws.append(r)
    
    return wb


def find_unique_values(jvc_df):
    try:
        unique_values = jvc_df["identifier"].unique()
    except:
        unique_values = jvc_df["sample"].unique()
    print(f"\nThe following samples were found in the dataset: {', '.join(map(str, unique_values))}")

    return unique_values