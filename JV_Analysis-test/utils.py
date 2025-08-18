"""
Utility Functions Module
Contains Excel export, file operations, and other utility functions.
Extracted from main.py for better organization.
"""

__author__ = "Edgar Nandayapa"
__institution__ = "Helmholtz-Zentrum Berlin"
__created__ = "August 2025"

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas as pd
import os


def save_full_data_frame(data):
    """
    Create and return an Excel workbook with the full dataframe.
    Simplified version that just creates a workbook without saving to file.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    
    # Add main data sheet
    ws = wb.create_sheet(title='All_data')
    for r in dataframe_to_rows(data, index=True, header=True):
        ws.append(r)
    
    return wb


def save_combined_excel_data(path, wb, data, filtered_info, var_x, name_y, var_y, other_df):
    """Save combined data to Excel workbook with multiple sheets"""
    trash, filters = filtered_info
    
    # Create sheet name based on variables
    sheet_title = f"{var_y}-by-{var_x}"

    # Check if the sheet already exists and remove it
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    # Insert header
    ws.append([f"Contents of boxplot for {var_y} by {var_x}"])
    ws.append([])  # Empty row

    # Process and append main data
    combined_data = data.copy()
    combined_data['_index'] = combined_data.groupby(var_x).cumcount()
    pivot_table = combined_data.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

    for r in dataframe_to_rows(pivot_table, index=True, header=True):
        ws.append(r)

    # Add statistical summary
    next_row = ws.max_row + 3
    ws.cell(row=next_row, column=1, value="Statistical summary")
    ws.append([])

    for r in dataframe_to_rows(other_df.T, index=True, header=True):
        ws.append(r)

    # Add filtered data section
    next_row = ws.max_row + 3
    ws.cell(row=next_row, column=1, value="This is the filtered data")
    ws.append([])

    if not trash.empty:
        combined_trash = trash.copy()
        combined_trash['_index'] = combined_trash.groupby(var_x).cumcount()
        pivot_table_trash = combined_trash.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

        for r in dataframe_to_rows(pivot_table_trash, index=True, header=True):
            ws.append(r)

    # Add filter information
    next_row = ws.max_row + 3
    filter_words = ["Only data within these limits is shown:"] + filters
    for cc, strings in enumerate(filter_words):
        ws.cell(row=next_row + cc, column=1, value=strings)

    return wb


def is_running_in_jupyter():
    """Check if code is running in Jupyter notebook"""
    try:
        from IPython import get_ipython
        return get_ipython() is not None
    except ImportError:
        return False


def natural_keys(text):
    """Natural sorting key function for alphanumeric strings"""
    import re
    def atoi(text):
        return int(text) if text.isdigit() else text
    return [atoi(c) for c in re.split(r'(\d+)', text)]


def create_new_results_folder(path):
    """Create a results folder if it doesn't exist"""
    folder_path = os.path.join(path, 'Results')
    try:
        os.makedirs(folder_path, exist_ok=True)
    except Exception as e:
        print(f"Warning: Could not create results folder: {e}")
        return path
    return folder_path


def clean_filename(filename):
    """Clean filename for safe saving"""
    import re
    # Remove invalid characters for filenames
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
    return filename


def export_to_csv(dataframes, path=""):
    """Export multiple dataframes to CSV files"""
    for name, df in dataframes.items():
        try:
            filename = os.path.join(path, f"export_{name}.csv")
            df.to_csv(filename, index=False)
            print(f"Exported {name} to {filename}")
        except Exception as e:
            print(f"Warning: Could not export {name} to CSV: {e}")


def validate_data_structure(data):
    """Validate that data structure contains required keys and is properly formatted"""
    required_keys = ['jvc', 'curves']
    
    if not isinstance(data, dict):
        return False, "Data must be a dictionary"
    
    for key in required_keys:
        if key not in data:
            return False, f"Missing required key: {key}"
        
        if not isinstance(data[key], pd.DataFrame):
            return False, f"Key '{key}' must contain a pandas DataFrame"
        
        if data[key].empty:
            return False, f"DataFrame '{key}' is empty"
    
    # Check for required columns in jvc data
    required_jvc_columns = ['PCE(%)', 'sample', 'cell']
    for col in required_jvc_columns:
        if col not in data['jvc'].columns:
            return False, f"Missing required column '{col}' in jvc data"
    
    return True, "Data structure is valid"


def get_file_size_mb(filepath):
    """Get file size in megabytes"""
    try:
        size_bytes = os.path.getsize(filepath)
        size_mb = size_bytes / (1024 * 1024)
        return round(size_mb, 2)
    except OSError:
        return 0


def create_download_link(content, filename, content_type='application/octet-stream'):
    """Create a download link for content (for use in Jupyter)"""
    import base64
    from IPython.display import HTML
    
    if isinstance(content, str):
        content = content.encode()
    
    b64 = base64.b64encode(content).decode()
    return HTML(f'''
    <a download="{filename}" 
       href="data:{content_type};base64,{b64}" 
       style="background-color: #4CAF50; color: white; padding: 10px 20px; 
              text-decoration: none; border-radius: 4px; display: inline-block;">
        Download {filename}
    </a>
    ''')